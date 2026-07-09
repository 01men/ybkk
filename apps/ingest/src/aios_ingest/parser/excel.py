"""aios_ingest.parser.excel —— Excel 解析器（openpyxl）。"""
from __future__ import annotations

import io
import logging
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook

from . import BaseParser, ParsedDocument

logger = logging.getLogger("aios_ingest.excel")


def _cell_to_value(v: Any) -> Any:
    """openpyxl cell → 友好值。"""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float, str)):
        return v
    return str(v)


def _infer_type(values: list[Any]) -> str:
    """简单类型推断。"""
    non_null = [v for v in values if v is not None]
    if not non_null:
        return "string"
    if all(isinstance(v, bool) for v in non_null):
        return "boolean"
    if all(isinstance(v, (int, float)) for v in non_null):
        return "number"
    if all(isinstance(v, (datetime, date)) for v in non_null):
        return "date"
    return "string"


class ExcelParser(BaseParser):
    """Excel 解析：第一行为表头，剩余为数据。"""

    kind = "excel"

    def parse(self, content: bytes, filename: str = "") -> ParsedDocument:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        rows: list[dict] = []
        sheet_names: list[str] = []
        all_tables: list[list[list[str]]] = []

        for sheet_name in wb.sheetnames:
            sheet_names.append(sheet_name)
            ws = wb[sheet_name]
            ws_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(ws_iter)
            except StopIteration:
                continue
            header = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header_row)]

            table: list[list[str]] = [header]
            for row in ws_iter:
                values = [_cell_to_value(v) for v in row]
                if all(v is None or v == "" for v in values):
                    continue
                rows.append(dict(zip(header, values, strict=False)))
                table.append([str(v) if v is not None else "" for v in values])
            all_tables.append(table)

        # 列类型推断
        columns_meta: list[dict] = []
        if rows:
            for h in rows[0].keys():
                col_values = [r.get(h) for r in rows]
                columns_meta.append({"name": h, "inferred_type": _infer_type(col_values)})

        # 拼装 text（供 LLM 用）
        text_lines = [f"Sheet: {s}, rows={len([r for r in rows])}" for s in sheet_names]
        text_lines.append("Columns:")
        for c in columns_meta:
            text_lines.append(f"  - {c['name']} ({c['inferred_type']})")
        text_lines.append("Sample rows (first 10):")
        for r in rows[:10]:
            text_lines.append("  " + " | ".join(f"{k}={v}" for k, v in r.items()))

        return ParsedDocument(
            kind=self.kind,
            rows=rows,
            text="\n".join(text_lines),
            tables=all_tables,
            metadata={
                "sheet_names": sheet_names,
                "columns": columns_meta,
                "row_count": len(rows),
                "filename": filename,
            },
        )


__all__ = ["ExcelParser"]
